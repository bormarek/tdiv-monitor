from flask import Flask, render_template, jsonify, request
import openpyxl
import yfinance as yf
import pandas as pd
import requests
import io
from datetime import datetime, timedelta
import os
import time
import sqlite3
import json
import threading
import re
from deep_translator import GoogleTranslator
from concurrent.futures import ThreadPoolExecutor, as_completed

app = Flask(__name__)

# ── Baza danych (cache) ───────────────────────────────────────────────────────

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'cache.db')

# L1 – pamięć (ultra-szybki odczyt), L2 – SQLite (persystencja)
_mem_cache: dict = {}  # {(namespace, key): (data, ts)}


def init_db():
    with sqlite3.connect(DB_PATH) as con:
        con.execute('''
            CREATE TABLE IF NOT EXISTS cache (
                namespace TEXT NOT NULL,
                key       TEXT NOT NULL,
                value     TEXT NOT NULL,
                ts        REAL NOT NULL,
                PRIMARY KEY (namespace, key)
            )
        ''')
        con.execute('PRAGMA journal_mode=WAL')  # równoległe odczyty podczas zapisu
        con.commit()


def cache_get(namespace: str, key: str, ttl: float):
    mem_key = (namespace, key)
    now = time.time()

    # L1 – pamięć
    entry = _mem_cache.get(mem_key)
    if entry is not None:
        data, ts = entry
        if now - ts < ttl:
            return data
        del _mem_cache[mem_key]

    # L2 – SQLite
    with sqlite3.connect(DB_PATH) as con:
        row = con.execute(
            'SELECT value, ts FROM cache WHERE namespace=? AND key=?',
            (namespace, key)
        ).fetchone()
    if row is None:
        return None
    value, ts = row
    if now - ts >= ttl:
        return None
    data = json.loads(value)
    _mem_cache[mem_key] = (data, ts)  # podgrzej L1
    return data


def cache_set(namespace: str, key: str, data) -> None:
    ts = time.time()
    _mem_cache[(namespace, key)] = (data, ts)
    with sqlite3.connect(DB_PATH) as con:
        con.execute(
            'INSERT OR REPLACE INTO cache (namespace, key, value, ts) VALUES (?, ?, ?, ?)',
            (namespace, key, json.dumps(data, default=str), ts)
        )
        con.commit()


def cache_delete(namespace: str, key: str) -> None:
    _mem_cache.pop((namespace, key), None)
    with sqlite3.connect(DB_PATH) as con:
        con.execute('DELETE FROM cache WHERE namespace=? AND key=?', (namespace, key))
        con.commit()


init_db()

# ── Konfiguracja funduszy ─────────────────────────────────────────────────────
# (cache warmer uruchamiany po zdefiniowaniu FUNDS i _build_fund_response)

FUNDS = {
    'tdiv': {
        'id':       'tdiv',
        'name':     'VanEck Morningstar Dividend Leaders ETF',
        'short':    'TDIV',
        'currency': 'USD',
        'url':      'https://www.vaneck.com/pl/pl/investments/dividend-etf/downloads/holdings/',
        'headers':  {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'pl-PL,pl;q=0.9',
        },
        'local_fallback': 'TDIV_nadzień_20260305.xlsx',
    },
    'swig80': {
        'id':       'swig80',
        'name':     'Beta ETF sWIG80TR',
        'short':    'sWIG80',
        'currency': 'PLN',
        'url':      'https://wp00102-api.agiofunds.pl/uploads/funds/Portfolio/archive/PRTF_Beta%20ETF%20sWIG80TR.xlsx',
        'headers':  {'User-Agent': 'Mozilla/5.0'},
        'local_fallback': None,
    },
    'mwig40': {
        'id':       'mwig40',
        'name':     'Beta ETF mWIG40TR',
        'short':    'mWIG40',
        'currency': 'PLN',
        'url':      'https://wp00102-api.agiofunds.pl/uploads/funds/Portfolio/archive/PRTF_Beta%20ETF%20mWIG40TR.xlsx',
        'headers':  {'User-Agent': 'Mozilla/5.0'},
        'local_fallback': None,
    },
    'wig20': {
        'id':       'wig20',
        'name':     'Beta ETF WIG20TR',
        'short':    'WIG20',
        'currency': 'PLN',
        'url':      'https://wp00102-api.agiofunds.pl/uploads/funds/Portfolio/archive/PRTF_Beta%20ETF%20WIG20TR.xlsx',
        'headers':  {'User-Agent': 'Mozilla/5.0'},
        'local_fallback': None,
    },
}

CACHE_TTL          = 900    # 15 minut
INFO_CACHE_TTL     = 86400  # 24 godziny
CALENDAR_CACHE_TTL = 3600   # 1 godzina
MARKET_CAL_TTL     = 3600   # 1 godzina

# ── Ticker helpers (TDIV / Bloomberg) ────────────────────────────────────────

TICKER_OVERRIDES = {
    'DBS SP':   'D05.SI',
    'OCBC SP':  'O39.SI',
    'UOB SP':   'U11.SI',
    'KEP SP':   'BN4.SI',
    'WIL SP':   'F34.SI',
    'NOVOB DC': 'NOVO-B.CO',
    'EDP PL':   'EDP.LS',
    'LUMI IT':  'LUMI.TA',
    'MZTF IT':  'MZTF.TA',
}

EXCHANGE_MAP = {
    'US': '',   'SW': '.SW',  'LN': '.L',   'FP': '.PA',
    'GR': '.DE','SM': '.MC',  'DC': '.CO',  'NA': '.AS',
    'IM': '.MI','AU': '.AX',  'HK': '.HK',  'JP': '.T',
    'BB': '.BR','SS': '.ST',  'NO': '.OL',  'SP': '.SI',
    'CN': '.TO','PW': '.WA',  'AV': '.VI',  'PL': '.LS',
    'IT': '.TA','SJ': '.JO',
    # Azja
    'KS': '.KS', 'KQ': '.KQ',            # Korea (KRX / KOSDAQ)
    'TT': '.TW',                          # Tajwan (TWSE)
    'TB': '.BK',                          # Tajlandia
    'MK': '.KL',                          # Malezja
    'IJ': '.JK',                          # Indonezja
    'IN': '.NS', 'IB': '.BO',            # Indie (NSE / BSE)
    'PM': '.PS',                          # Filipiny
    # Ameryka Łac.
    'BZ': '.SA',                          # Brazylia
    'MM': '.MX',                          # Meksyk
}

# Pełne nazwy giełd → sufiks yfinance (dla generycznego parsera)
EXCHANGE_NAME_MAP = {
    'korea stock exchange': '.KS', 'krx': '.KS', 'kospi': '.KS',
    'kosdaq': '.KQ',
    'taiwan stock exchange': '.TW', 'twse': '.TW',
    'taiwan otc': '.TWO', 'taipei exchange': '.TWO',
    'tokyo stock exchange': '.T', 'tse': '.T', 'jpx': '.T',
    'hong kong stock exchange': '.HK', 'hkex': '.HK',
    'shanghai stock exchange': '.SS', 'sse': '.SS',
    'shenzhen stock exchange': '.SZ', 'szse': '.SZ',
    'national stock exchange of india': '.NS', 'nse india': '.NS',
    'bombay stock exchange': '.BO', 'bse india': '.BO',
    'singapore exchange': '.SI', 'sgx': '.SI',
    'australian securities exchange': '.AX', 'asx': '.AX',
    'bursa malaysia': '.KL',
    'stock exchange of thailand': '.BK', 'set': '.BK',
    'indonesia stock exchange': '.JK', 'idx': '.JK',
    'london stock exchange': '.L', 'lse': '.L',
    'euronext amsterdam': '.AS', 'euronext paris': '.PA',
    'euronext brussels': '.BR', 'euronext lisbon': '.LS',
    'xetra': '.DE', 'deutsche boerse': '.DE',
    'six swiss exchange': '.SW', 'swiss exchange': '.SW',
    'oslo bors': '.OL', 'oslo stock exchange': '.OL',
    'nasdaq stockholm': '.ST', 'nasdaq helsinki': '.HE',
    'nasdaq copenhagen': '.CO',
    'borsa italiana': '.MI',
    'bolsa de madrid': '.MC', 'bmex': '.MC',
    'vienna stock exchange': '.VI',
    'warsaw stock exchange': '.WA', 'gpw': '.WA',
    'johannesburg stock exchange': '.JO', 'jse': '.JO',
    'tel aviv stock exchange': '.TA', 'tase': '.TA',
    'toronto stock exchange': '.TO', 'tsx': '.TO',
    'new york stock exchange': '', 'nyse': '', 'nasdaq': '', 'nyse arca': '',
    # Dwuliterowe kody krajów (np. kolumna "Stock market" w plikach Vanguard)
    'us': '', 'gb': '.L', 'de': '.DE', 'fr': '.PA', 'ch': '.SW',
    'nl': '.AS', 'be': '.BR', 'pt': '.LS', 'es': '.MC', 'it': '.MI',
    'se': '.ST', 'no': '.OL', 'dk': '.CO', 'fi': '.HE', 'at': '.VI',
    'pl': '.WA', 'au': '.AX', 'sg': '.SI', 'hk': '.HK', 'jp': '.T',
    'kr': '.KS', 'tw': '.TW', 'cn': '.SS', 'in': '.NS', 'my': '.KL',
    'th': '.BK', 'id': '.JK', 'il': '.TA', 'za': '.JO', 'br': '.SA',
    'ca': '.TO', 'mx': '.MX',
}

# Prefiks kraju ISIN → sufiks yfinance (ostatnia szansa)
ISIN_COUNTRY_TO_EXCHANGE = {
    'KR': '.KS', 'JP': '.T',  'HK': '.HK', 'TW': '.TW',
    'CN': '.SS', 'IN': '.NS', 'SG': '.SI', 'AU': '.AX',
    'MY': '.KL', 'TH': '.BK', 'ID': '.JK', 'IL': '.TA',
    'BR': '.SA', 'MX': '.MX', 'ZA': '.JO',
}


def bloomberg_to_yfinance(ticker_str):
    key = ticker_str.strip()
    if key in TICKER_OVERRIDES:
        return TICKER_OVERRIDES[key]
    parts = key.split()
    if len(parts) < 2:
        return None
    symbol, exchange = parts[0], parts[-1]
    if exchange == '--' or symbol == '--':
        return None
    if '/' in symbol:
        symbol = symbol.rstrip('/') if symbol.endswith('/') else symbol.replace('/', '-')
    if exchange == 'HK' and symbol.isdigit():
        symbol = symbol.zfill(4)
    return symbol + EXCHANGE_MAP.get(exchange, '')


# ── Parsery Excel ─────────────────────────────────────────────────────────────

def parse_tdiv(wb):
    ws = wb.active
    holdings = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None or not isinstance(row[0], int):
            continue
        name, ticker_raw, weight_str = row[1], row[2], row[6]
        if not name or not ticker_raw:
            continue
        yf_ticker = bloomberg_to_yfinance(str(ticker_raw))
        if not yf_ticker:
            continue
        holdings.append({
            'number':     row[0],
            'name':       name,
            'ticker_raw': str(ticker_raw).strip(),
            'ticker':     yf_ticker,
            'weight':     weight_str,
            'sector':     '',
        })
    return holdings


def parse_swig80(wb):
    ws = wb.active
    holdings = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] == 'Lp.':
                header_found = True
            continue
        if row[0] is None or not isinstance(row[0], int):
            continue
        name    = row[1]
        isin    = row[2]
        sector  = row[3] or ''
        weight  = row[6]  # ułamek dziesiętny, np. 0.044
        if not name or not isin or isin == '---' or not isin.replace('-', '').isalnum() or str(isin).startswith('PL0'):
            continue
        weight_str = f"{weight * 100:.2f}%".replace('.', ',') if weight else None
        yf_ticker = POLISH_ISIN_OVERRIDES.get(isin, isin)
        holdings.append({
            'number':     row[0],
            'name':       name,
            'ticker_raw': isin,
            'ticker':     yf_ticker,
            'weight':     weight_str,
            'sector':     sector,
        })
    return holdings


# ISINy które yfinance rozwiązuje na giełdę zagraniczną zamiast GPW (.WA)
POLISH_ISIN_OVERRIDES = {
    # Zagraniczne ISINy spółek notowanych na GPW
    'LU2237380790': 'ALE.WA',    # Allegro.eu
    'AU0000198939': 'GRX.WA',    # GreenX Metals
    # Polskie ISINy trafiające na Stuttgart (EUR) zamiast GPW (PLN)
    'PLAMBRA00013': 'AMB.WA',    # Ambra
    'PLBMDLB00018': 'BIO.WA',    # Biomed-Lublin
    'PLBUDMX00013': 'BDX.WA',    # Budimex
    'PLCMPLD00016': 'SGN.WA',    # Sygnity
    'PLCPTRT00014': 'CTX.WA',    # Captor Therapeutics
    'PLDADEL00013': 'DAD.WA',    # Dadelo
    'PLECHPS00019': 'ECH.WA',    # Echo Investment
    'PLERBUD00012': 'ERB.WA',    # Erbud
    'PLGPW0000017': 'GPW.WA',    # Giełda Papierów Wartościowych
    'PLLVTSF00010': 'TXT.WA',    # Text SA
    'PLOPNPL00013': 'OPN.WA',    # Oponeo.pl
    'PLSTLEX00019': 'STX.WA',    # Stalexport Autostrady
    'PLTOYA000011': 'TOA.WA',    # TOYA
}

PARSERS = {'tdiv': parse_tdiv, 'swig80': parse_swig80, 'mwig40': parse_swig80, 'wig20': parse_swig80}


def load_holdings(fund_id):
    fund = FUNDS[fund_id]
    source = 'live'
    try:
        r = requests.get(fund['url'], headers=fund['headers'], timeout=30)
        r.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
    except Exception as e:
        source = f'local (błąd pobierania: {e})'
        if fund['local_fallback']:
            path = os.path.join(os.path.dirname(os.path.abspath(__file__)), fund['local_fallback'])
            wb = openpyxl.load_workbook(path)
        else:
            return [], source
    return PARSERS[fund_id](wb), source


# ── Analiza cenowa ────────────────────────────────────────────────────────────

def analyze_series(series):
    result = {
        'price': None, 'daily_change': None, 'daily_change_pct': None,
        'daily_trend': 'unknown', 'ma5': None, 'ma5_trend': 'unknown',
    }
    if series is None or len(series) < 2:
        return result
    last, prev = float(series.iloc[-1]), float(series.iloc[-2])
    result['price']            = round(last, 2)
    result['daily_change']     = round(last - prev, 4)
    result['daily_change_pct'] = round((last - prev) / prev * 100, 2)
    result['daily_trend']      = 'up' if last >= prev else 'down'
    if len(series) >= 6:
        ma5_cur  = float(series.iloc[-5:].mean())
        ma5_prev = float(series.iloc[-6:-1].mean())
        result['ma5']       = round(ma5_cur, 2)
        result['ma5_trend'] = 'up' if ma5_cur > ma5_prev else 'down'
    elif len(series) >= 5:
        result['ma5'] = round(float(series.iloc[-5:].mean()), 2)
    return result


# ── Generyczny parser plików (Excel / CSV) ────────────────────────────────────

def _isin_to_local_ticker(isin: str) -> str:
    """Wyciąga lokalny ticker giełdowy z ISIN dla znanych formatów."""
    country = isin[:2]
    try:
        if country == 'KR':
            # KR7XXXXXX000 → ISIN[3:9] = 6-cyfrowy kod KRX
            return isin[3:9]
        if country == 'TW':
            # TW000XXXX00Y → int(ISIN[2:9]) daje 4-cyfrowy kod TWSE
            return str(int(isin[2:9]))
    except Exception:
        pass
    return ''


def _download_prices_chunked(tickers: list, start: str, end: str, chunk_size: int = 100) -> pd.DataFrame:
    """Pobiera ceny zamknięcia w partiach, łączy wyniki."""
    all_close = pd.DataFrame()
    for i in range(0, len(tickers), chunk_size):
        chunk = tickers[i:i + chunk_size]
        try:
            raw = yf.download(chunk, start=start, end=end,
                              auto_adjust=True, progress=False, threads=True)
            if raw.empty:
                continue
            if isinstance(raw.columns, pd.MultiIndex):
                close = raw['Close']
            else:
                close = raw[['Close']]
                close.columns = chunk
            all_close = close if all_close.empty else all_close.join(close, how='outer')
        except Exception:
            continue
    return all_close


_COL_WEIGHT      = re.compile(r'weight|alloc|% of|waga|udzia', re.I)
_COL_NAME        = re.compile(r'^(name|security|holding|company|issuer|instrument|nazwa|emitent|asset name|security desc)', re.I)
_COL_TICKER      = re.compile(r'^(ticker|symbol)$', re.I)
_COL_ISIN        = re.compile(r'isin', re.I)
_COL_EXCHANGE    = re.compile(r'^(exchange|market|giełda|rynek)', re.I)
_COL_ASSET_CLASS = re.compile(r'^(asset.?class|klasa aktyw|asset type)', re.I)
_EQUITY_CLASS    = re.compile(r'^equity|stock|akcj', re.I)
_SKIP_NAMES      = re.compile(
    r'^(cash|total|other assets|other|razem|gotówka|xgld|xtreasury|money market'
    r'|fixed income|bond|treasury|derivative|futures|forward|option|swap'
    r'|blackrock cash|ishares cash|ishares liquidity|liquidity fund'
    r'|n\/a|\-+|\.+)$',
    re.I
)


def parse_generic_file(file_bytes: bytes, filename: str) -> list:
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    # Odczyt do surowego DataFrame (brak nagłówka)
    df_raw = None
    if ext == 'csv':
        for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1250'):
            for sep in (',', ';', '\t'):
                try:
                    df = pd.read_csv(io.BytesIO(file_bytes), header=None,
                                     encoding=enc, sep=sep, on_bad_lines='skip', dtype=str)
                    if df.shape[1] >= 3:
                        df_raw = df
                        break
                except Exception:
                    continue
            if df_raw is not None:
                break
    else:
        try:
            df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
        except Exception as e:
            raise ValueError(f'Błąd odczytu pliku Excel: {e}')

    if df_raw is None or df_raw.shape[1] < 2:
        raise ValueError('Nie można odczytać pliku. Sprawdź format (xlsx/xls/csv).')

    df_raw = df_raw.fillna('')

    # Wykryj wiersz nagłówkowy
    header_idx = None
    for i, row in df_raw.iterrows():
        cells = [str(c).strip() for c in row if str(c).strip()]
        if len(cells) < 2:
            continue
        score = sum(1 for c in cells if (
            _COL_WEIGHT.search(c) or _COL_NAME.match(c) or
            _COL_TICKER.match(c) or _COL_ISIN.search(c)
        ))
        if score >= 1:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError('Nie można wykryć nagłówka tabeli. Sprawdź format pliku.')

    headers = [str(c).strip() or f'col_{j}' for j, c in enumerate(df_raw.iloc[header_idx])]
    data = df_raw.iloc[header_idx + 1:].reset_index(drop=True)
    data.columns = headers

    # Mapowanie kolumn
    name_col = ticker_col = isin_col = weight_col = exchange_col = asset_class_col = None
    for h in headers:
        if name_col        is None and _COL_NAME.match(h):        name_col        = h
        if ticker_col      is None and _COL_TICKER.match(h):      ticker_col      = h
        if isin_col        is None and _COL_ISIN.search(h):       isin_col        = h
        if weight_col      is None and _COL_WEIGHT.search(h):     weight_col      = h
        if exchange_col    is None and _COL_EXCHANGE.match(h):    exchange_col    = h
        if asset_class_col is None and _COL_ASSET_CLASS.match(h): asset_class_col = h

    if name_col is None:
        name_col = headers[0]  # fallback: pierwsza kolumna
    if ticker_col is None and isin_col is None:
        raise ValueError('Nie znaleziono kolumny z tickerem ani ISIN.')

    # Skala wag (ułamek 0-1 vs procent 0-100)
    raw_weights = []
    if weight_col:
        for v in data[weight_col]:
            try:
                w = float(str(v).replace(',', '.').replace('%', '').strip())
                if w > 0:
                    raw_weights.append(w)
            except Exception:
                pass
    weight_scale = 100.0 if raw_weights and max(raw_weights) <= 1.5 else 1.0

    holdings = []
    for _, row in data.iterrows():
        name = str(row.get(name_col, '')).strip()
        if not name or name in ('nan', 'None') or _SKIP_NAMES.match(name):
            continue

        # Pomiń wiersze nie będące akcjami (gotówka, obligacje, derywaty…)
        if asset_class_col:
            ac = str(row.get(asset_class_col, '')).strip()
            if ac and not _EQUITY_CLASS.match(ac):
                continue

        weight_str = None
        if weight_col:
            try:
                w = float(str(row.get(weight_col, '')).replace(',', '.').replace('%', '').strip())
                if w > 0:
                    weight_str = f'{w * weight_scale:.2f}%'.replace('.', ',')
            except Exception:
                pass

        ticker_raw = ''
        yf_ticker  = None

        # Nazwa giełdy → sufiks yfinance
        exch_suffix = ''
        if exchange_col:
            exch_name = str(row.get(exchange_col, '')).strip().lower()
            exch_suffix = EXCHANGE_NAME_MAP.get(exch_name, '')

        if ticker_col:
            t = str(row.get(ticker_col, '')).strip()
            if t and t not in ('nan', '-', '', '—'):
                ticker_raw = t
                if ' ' in t and len(t.split()[-1]) == 2:
                    # Bloomberg format: "012450 KS"
                    yf_ticker = bloomberg_to_yfinance(t)
                elif exch_suffix:
                    # Lokalny ticker + znana giełda: "012450" + ".KS"
                    yf_ticker = t + exch_suffix
                else:
                    yf_ticker = t

        if isin_col and not yf_ticker:
            isin = str(row.get(isin_col, '')).strip()
            if isin and len(isin) == 12 and isin[:2].isalpha():
                ticker_raw = ticker_raw or isin
                if isin in POLISH_ISIN_OVERRIDES:
                    yf_ticker = POLISH_ISIN_OVERRIDES[isin]
                else:
                    country = isin[:2]
                    suffix  = exch_suffix or ISIN_COUNTRY_TO_EXCHANGE.get(country, '')
                    local   = _isin_to_local_ticker(isin)
                    if local and suffix:
                        # Znamy format ISIN i giełdę: "005930" + ".KS"
                        yf_ticker = local + suffix
                    elif suffix == '':
                        # Rynek US lub inny bez sufiksu → przekaż ISIN
                        # (yfinance.Ticker obsługuje ISINy, download może nie)
                        yf_ticker = isin
                    else:
                        # Nieznany format lokalny, ale znamy sufiks → użyj ISIN
                        yf_ticker = isin

        if not yf_ticker:
            continue

        holdings.append({
            'number':     len(holdings) + 1,
            'name':       name,
            'ticker_raw': ticker_raw,
            'ticker':     yf_ticker,
            'weight':     weight_str,
            'sector':     '',
        })

    if not holdings:
        raise ValueError('Nie wykryto żadnych pozycji. Sprawdź format pliku.')
    return holdings


# ── Upload sesyjny ─────────────────────────────────────────────────────────────

_upload_session = None


# ── Endpointy ─────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html', funds=list(FUNDS.values()))


@app.route('/api/funds')
def get_funds():
    return jsonify(list(FUNDS.values()))


def _build_fund_response(fund_id: str) -> dict:
    """Pobiera dane funduszu z zewnątrz i zwraca gotowy słownik odpowiedzi."""
    holdings, holdings_source = load_holdings(fund_id)
    tickers = [h['ticker'] for h in holdings]

    end_date   = datetime.now()
    start_date = end_date - timedelta(days=25)
    raw = yf.download(
        tickers,
        start=start_date.strftime('%Y-%m-%d'),
        end=end_date.strftime('%Y-%m-%d'),
        auto_adjust=True, progress=False, threads=True,
    )

    if isinstance(raw.columns, pd.MultiIndex):
        close = raw['Close']
    else:
        close = raw[['Close']]
        close.columns = tickers

    missing = [t for t in tickers if t not in close.columns]
    if missing:
        try:
            raw2 = yf.download(
                missing,
                start=start_date.strftime('%Y-%m-%d'),
                end=end_date.strftime('%Y-%m-%d'),
                auto_adjust=True, progress=False, threads=True,
            )
            if not raw2.empty:
                close2 = raw2['Close'] if isinstance(raw2.columns, pd.MultiIndex) else raw2[['Close']]
                if not isinstance(raw2.columns, pd.MultiIndex):
                    close2.columns = missing
                close = close.join(close2, how='outer')
        except Exception:
            pass

    results = []
    for h in holdings:
        entry = {
            'number':    h['number'],
            'name':      h['name'],
            'ticker':    h['ticker_raw'],
            'yf_ticker': h['ticker'],
            'weight':    h['weight'],
            'sector':    h.get('sector', ''),
        }
        try:
            series = close[h['ticker']].dropna() if h['ticker'] in close.columns else None
            entry.update(analyze_series(series))
        except Exception as e:
            entry.update(analyze_series(None))
            entry['error'] = str(e)
        results.append(entry)

    up_count   = sum(1 for r in results if r.get('daily_trend') == 'up')
    down_count = sum(1 for r in results if r.get('daily_trend') == 'down')

    return {
        'data':            results,
        'updated_at':      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'holdings_source': holdings_source,
        'fund':            FUNDS[fund_id],
        'summary': {
            'up':      up_count,
            'down':    down_count,
            'unknown': len(results) - up_count - down_count,
        },
    }


def _warm_fund_cache():
    """Odświeża cache wszystkich funduszy w tle (pomija świeże wpisy)."""
    for fund_id in FUNDS:
        if cache_get('fund_data', fund_id, CACHE_TTL) is not None:
            continue
        try:
            response = _build_fund_response(fund_id)
            cache_set('fund_data', fund_id, response)
        except Exception:
            pass


def _start_cache_warmer():
    """Uruchamia wątek podgrzewający cache; odświeża co CACHE_TTL - 60 s."""
    def loop():
        _warm_fund_cache()           # pierwsze uruchomienie przy starcie
        while True:
            time.sleep(CACHE_TTL - 60)
            _warm_fund_cache()
    t = threading.Thread(target=loop, daemon=True, name='cache-warmer')
    t.start()


@app.route('/api/data')
def get_data():
    fund_id = request.args.get('fund', 'tdiv')
    if fund_id not in FUNDS:
        return jsonify({'error': f'Nieznany fundusz: {fund_id}'}), 400

    cached = cache_get('fund_data', fund_id, CACHE_TTL)
    if cached is not None:
        return jsonify(cached)

    try:
        response = _build_fund_response(fund_id)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    cache_set('fund_data', fund_id, response)
    return jsonify(response)


@app.route('/api/info/<path:ticker>')
def get_info(ticker):
    cached = cache_get('company_info', ticker, INFO_CACHE_TTL)
    if cached is not None:
        return jsonify(cached)
    try:
        info = yf.Ticker(ticker).info
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    description_en = info.get('longBusinessSummary', '')
    description_pl = ''
    if description_en:
        try:
            description_pl = GoogleTranslator(source='en', target='pl').translate(description_en)
        except Exception:
            description_pl = description_en

    data = {
        'name':        info.get('longName') or info.get('shortName', ''),
        'sector':      info.get('sector', ''),
        'industry':    info.get('industry', ''),
        'country':     info.get('country', ''),
        'description': description_pl,
        'website':     info.get('website', ''),
        'employees':   info.get('fullTimeEmployees'),
    }
    if data['description']:
        cache_set('company_info', ticker, data)
    return jsonify(data)


def _safe_float(v):
    try:
        f = float(v)
        return None if pd.isna(f) else round(f, 4)
    except Exception:
        return None


@app.route('/api/calendar/<path:ticker>')
def get_calendar(ticker):
    cached = cache_get('ticker_calendar', ticker, CALENDAR_CACHE_TTL)
    if cached is not None:
        return jsonify(cached)

    t = yf.Ticker(ticker)
    result = {'earnings': [], 'dividends': [], 'news': []}

    try:
        ed = t.get_earnings_dates(limit=8)
        if ed is not None and not ed.empty:
            for date, row in ed.iterrows():
                reported = _safe_float(row.get('Reported EPS'))
                estimate = _safe_float(row.get('EPS Estimate'))
                surprise = _safe_float(row.get('Surprise(%)'))
                if surprise is not None:
                    surprise = round(surprise, 2)
                result['earnings'].append({
                    'date':     date.strftime('%Y-%m-%d'),
                    'reported': reported,
                    'estimate': estimate,
                    'surprise': surprise,
                })
    except Exception:
        pass

    try:
        divs = t.dividends
        if divs is not None and not divs.empty:
            recent = divs.tail(6).iloc[::-1]
            result['dividends'] = [
                {'date': str(d.date()), 'amount': round(float(v), 4)}
                for d, v in zip(recent.index, recent.values)
            ]
    except Exception:
        pass

    try:
        news_list = t.news or []
        for n in news_list[:6]:
            c = n.get('content', n)  # nowy format: {id, content:{...}}, stary: płaski
            title = c.get('title', '')
            publisher = (c.get('provider') or {}).get('displayName', '') or c.get('publisher', '')
            url = ((c.get('clickThroughUrl') or {}).get('url', '')
                   or (c.get('canonicalUrl') or {}).get('url', '')
                   or c.get('link', ''))
            pub_date = c.get('pubDate', '') or c.get('displayTime', '')
            if pub_date:
                date_str = pub_date[:10]  # ISO format YYYY-MM-DD
            else:
                ts = c.get('providerPublishTime', 0)
                date_str = datetime.fromtimestamp(ts).strftime('%Y-%m-%d') if ts else ''
            if title:
                result['news'].append({
                    'title':     title,
                    'publisher': publisher,
                    'url':       url,
                    'date':      date_str,
                })
    except Exception:
        pass

    # Tłumaczenie newsów dla polskich spółek
    if ticker.upper().endswith('.WA') and result['news']:
        try:
            SEP = '\n<|||>\n'
            combined = SEP.join(n['title'] for n in result['news'])
            translated = GoogleTranslator(source='en', target='pl').translate(combined)
            parts = translated.split(SEP)
            for i, n in enumerate(result['news']):
                if i < len(parts) and parts[i].strip():
                    n['title'] = parts[i].strip()
        except Exception:
            pass

    if result['earnings'] or result['dividends'] or result['news']:
        cache_set('ticker_calendar', ticker, result)

    return jsonify(result)


@app.route('/api/market-calendar')
def get_market_calendar():
    fund_id = request.args.get('fund', 'swig80')
    if fund_id not in FUNDS:
        return jsonify({'error': f'Nieznany fundusz: {fund_id}'}), 400

    cached = cache_get('market_calendar', fund_id, MARKET_CAL_TTL)
    if cached is not None:
        return jsonify(cached)

    holdings, _ = load_holdings(fund_id)
    today_str = datetime.now().strftime('%Y-%m-%d')

    def fetch_one(holding):
        ticker = holding['ticker']
        name   = holding['name']
        events = []
        try:
            t = yf.Ticker(ticker)
            # Nadchodzące raporty wynikowe
            try:
                ed = t.get_earnings_dates(limit=6)
                if ed is not None and not ed.empty:
                    for date, row in ed.iterrows():
                        reported = _safe_float(row.get('Reported EPS'))
                        estimate = _safe_float(row.get('EPS Estimate'))
                        surprise = _safe_float(row.get('Surprise(%)'))
                        if surprise is not None:
                            surprise = round(surprise, 2)
                        events.append({
                            'type':     'earnings',
                            'date':     date.strftime('%Y-%m-%d'),
                            'name':     name,
                            'ticker':   ticker,
                            'estimate': estimate,
                            'reported': reported,
                            'surprise': surprise,
                        })
            except Exception:
                pass
            # Dywidendy
            try:
                divs = t.dividends
                if divs is not None and not divs.empty:
                    for d, v in zip(divs.tail(6).index, divs.tail(6).values):
                        events.append({
                            'type':   'dividend',
                            'date':   str(d.date()),
                            'name':   name,
                            'ticker': ticker,
                            'amount': round(float(v), 4),
                        })
            except Exception:
                pass
        except Exception:
            pass
        return events

    all_events = []
    with ThreadPoolExecutor(max_workers=12) as executor:
        futures = {executor.submit(fetch_one, h): h for h in holdings}
        for f in as_completed(futures):
            all_events.extend(f.result())

    upcoming = sorted(
        [e for e in all_events if e['date'] >= today_str],
        key=lambda x: x['date'],
    )
    recent = sorted(
        [e for e in all_events if e['date'] < today_str],
        key=lambda x: x['date'],
        reverse=True,
    )

    result = {
        'upcoming': upcoming[:50],
        'recent':   recent[:40],
        'fund':     FUNDS[fund_id],
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    cache_set('market_calendar', fund_id, result)
    return jsonify(result)


@app.route('/api/chart/<path:ticker>')
def get_chart(ticker):
    # Obsługiwane okresy: 1d, 1wk, 1mo (domyślny), 1y
    period = request.args.get('period', '1mo')
    PERIOD_CONFIG = {
        '1d':  {'period': '1d',  'interval': '5m'},
        '1wk': {'period': '5d',  'interval': '1h'},
        '1mo': {'period': '6mo', 'interval': '1d'},
        '1y':  {'period': '1y',  'interval': '1d'},
    }
    if period not in PERIOD_CONFIG:
        period = '1mo'
    cfg = PERIOD_CONFIG[period]
    intraday = period in ('1d', '1wk')
    try:
        raw = yf.download(
            ticker,
            period=cfg['period'],
            interval=cfg['interval'],
            auto_adjust=True, progress=False,
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    if raw.empty:
        return jsonify({'error': 'Brak danych dla tego tickera'}), 404

    def col(name):
        return raw[name].iloc[:, 0] if isinstance(raw.columns, pd.MultiIndex) else raw[name]

    close  = col('Close')
    open_  = col('Open')
    high   = col('High')
    low    = col('Low')
    volume = col('Volume')

    # Wyrównaj indeks do dni gdy close nie jest NaN
    idx   = close.dropna().index
    close = close.loc[idx]
    open_ = open_.loc[idx]
    high  = high.loc[idx]
    low   = low.loc[idx]
    volume = volume.loc[idx]

    ma5  = close.rolling(5).mean()
    ma20 = close.rolling(20).mean()
    fmt  = lambda v: round(float(v), 2) if not pd.isna(v) else None
    fmtv = lambda v: int(v) if not pd.isna(v) else None

    def _calc_rsi(series, period=14):
        delta = series.diff()
        gain = delta.where(delta > 0, 0.0)
        loss = -delta.where(delta < 0, 0.0)
        avg_gain = gain.ewm(com=period-1, min_periods=period).mean()
        avg_loss = loss.ewm(com=period-1, min_periods=period).mean()
        rs = avg_gain / avg_loss.replace(0, float('inf'))
        return 100 - (100 / (1 + rs))

    def _calc_macd(series, fast=12, slow=26, signal=9):
        ema_fast = series.ewm(span=fast, adjust=False).mean()
        ema_slow = series.ewm(span=slow, adjust=False).mean()
        macd = ema_fast - ema_slow
        sig = macd.ewm(span=signal, adjust=False).mean()
        return macd, sig, macd - sig

    def _calc_bb(series, period=20, std=2):
        mid = series.rolling(period).mean()
        sd = series.rolling(period).std()
        return mid + std*sd, mid, mid - std*sd

    rsi = _calc_rsi(close)
    bb_upper, bb_middle, bb_lower = _calc_bb(close)
    macd, macd_signal, macd_hist = _calc_macd(close)

    # Trend
    if len(ma20.dropna()) >= 21:
        ma20_now  = ma20.dropna().iloc[-1]
        ma20_prev = ma20.dropna().iloc[-21] if len(ma20.dropna()) >= 21 else ma20.dropna().iloc[0]
        price_now = close.iloc[-1]
        ma20_rising = ma20_now > ma20_prev
        price_above = price_now > ma20_now
        if ma20_rising and price_above:
            trend = 'bullish'
        elif not ma20_rising and not price_above:
            trend = 'bearish'
        else:
            trend = 'sideways'
    else:
        trend = 'unknown'

    # Format dat: intraday → Unix timestamp (sekundy UTC) dla lightweight-charts, daily → YYYY-MM-DD
    if intraday:
        date_strs = []
        for d in idx:
            if hasattr(d, 'timestamp'):
                date_strs.append(int(d.timestamp()))
            else:
                date_strs.append(str(d))
    else:
        date_strs = [d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d)[:10] for d in idx]

    return jsonify({
        'ticker':      ticker,
        'period':      period,
        'intraday':    intraday,
        'dates':       date_strs,
        'prices':      [fmt(p) for p in close.values],
        'open':        [fmt(v) for v in open_.values],
        'high':        [fmt(v) for v in high.values],
        'low':         [fmt(v) for v in low.values],
        'volume':      [fmtv(v) for v in volume.values],
        'ma5':         [fmt(v) for v in ma5],
        'ma20':        [fmt(v) for v in ma20],
        'rsi':         [fmt(v) for v in rsi],
        'bb_upper':    [fmt(v) for v in bb_upper],
        'bb_middle':   [fmt(v) for v in bb_middle],
        'bb_lower':    [fmt(v) for v in bb_lower],
        'macd':        [fmt(v) for v in macd],
        'macd_signal': [fmt(v) for v in macd_signal],
        'macd_hist':   [fmt(v) for v in macd_hist],
        'trend':       trend,
    })


@app.route('/api/upload', methods=['POST'])
def upload_file():
    global _upload_session
    if 'file' not in request.files:
        return jsonify({'error': 'Brak pliku w żądaniu'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'Nie wybrano pliku'}), 400
    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in ('xlsx', 'xls', 'csv'):
        return jsonify({'error': 'Obsługiwane formaty: xlsx, xls, csv'}), 400

    try:
        holdings = parse_generic_file(f.read(), f.filename)
    except ValueError as e:
        return jsonify({'error': str(e)}), 422

    # Rozwiąż US ISINy → symbole (yfinance.Ticker obsługuje ISINy,
    # ale yf.download() już nie — robimy batch-resolve w tle)
    def _resolve_isin(h):
        t = h['ticker']
        if len(t) == 12 and t[:2].isalpha() and not t[2].isalpha():
            try:
                sym = yf.Ticker(t).fast_info.get('symbol') or ''
                if sym:
                    return h['number'], sym
            except Exception:
                pass
        return h['number'], t

    us_isin_holdings = [h for h in holdings
                        if len(h['ticker']) == 12 and h['ticker'][:2] == 'US']
    if us_isin_holdings:
        with ThreadPoolExecutor(max_workers=20) as ex:
            resolved = dict(ex.map(_resolve_isin, us_isin_holdings))
        for h in holdings:
            if h['number'] in resolved:
                h['ticker'] = resolved[h['number']]

    tickers    = [h['ticker'] for h in holdings]
    end_date   = datetime.now()
    start_date = end_date - timedelta(days=25)
    try:
        close = _download_prices_chunked(
            tickers,
            start_date.strftime('%Y-%m-%d'),
            end_date.strftime('%Y-%m-%d'),
        )
    except Exception as e:
        return jsonify({'error': f'Błąd pobierania cen: {e}'}), 500

    results = []
    for h in holdings:
        entry = {
            'number':    h['number'],
            'name':      h['name'],
            'ticker':    h['ticker_raw'],
            'yf_ticker': h['ticker'],
            'weight':    h['weight'],
            'sector':    '',
        }
        try:
            series = close[h['ticker']].dropna() if h['ticker'] in close.columns else None
            entry.update(analyze_series(series))
        except Exception:
            entry.update(analyze_series(None))
        results.append(entry)

    up_count   = sum(1 for r in results if r.get('daily_trend') == 'up')
    down_count = sum(1 for r in results if r.get('daily_trend') == 'down')

    _upload_session = {
        'data':            results,
        'updated_at':      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'holdings_source': f.filename,
        'fund': {'id': 'upload', 'name': f.filename, 'short': 'Własny', 'currency': '—'},
        'summary': {'up': up_count, 'down': down_count,
                    'unknown': len(results) - up_count - down_count},
    }
    return jsonify(_upload_session)


@app.route('/api/refresh')
def refresh_cache():
    fund_id = request.args.get('fund', 'tdiv')
    cache_delete('fund_data', fund_id)
    return jsonify({'status': 'ok'})


_start_cache_warmer()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=False, threaded=True)
